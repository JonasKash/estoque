import os
import pandas as pd
import glob
from flask import Flask, request, jsonify
from flask_cors import CORS
import logging
import unicodedata
import re
from openpyxl import load_workbook

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})

# Função para normalizar colunas

def normalize_column(col):
    if pd.isna(col):
        return ""
    col = str(col).strip().lower()
    col = ''.join(c for c in unicodedata.normalize('NFD', col) if unicodedata.category(c) != 'Mn')
    if col in ['cod. item', 'cod item', 'codigo item', 'cod_item']:
        return 'numero da peca'
    if col in ['descric?o', 'descrição', 'descricao', 'descri']:
        return 'descricao'
    if col in ['estoque', 'quantidade', 'qtd']:
        return 'quantidade'
    if col in ['locacao', 'locação', 'localizacao', 'localização', 'local']:
        return 'localizacao'
    col = col.replace('numero', 'num').replace('nº', 'num').replace('peca', 'peca').replace('peça', 'peca')
    col = col.replace('descrição', 'descricao').replace('descri', 'descricao')
    col = col.replace('qtd', 'quantidade').replace('quant.', 'quantidade').replace('quant', 'quantidade')
    col = col.replace('localizacao', 'localizacao').replace('localização', 'localizacao')
    return col

def extrair_data_planilha(nome_arquivo):
    match = re.search(r'(\d{6})', nome_arquivo)
    if match:
        data_str = match.group(1)
        dia = data_str[:2]
        mes = data_str[2:4]
        ano = '20' + data_str[4:]
        return f"{dia}/{mes}/{ano}"
    return nome_arquivo

def load_data_from_file(file_path, file_name):
    try:
        df = pd.read_excel(file_path, engine='openpyxl', header=2)
    except Exception:
        try:
            df = pd.read_excel(file_path, engine=None, header=2)
        except Exception:
            try:
                workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
                sheet = workbook.active
                data_rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
                if data_rows:
                    df = pd.DataFrame(data_rows[1:], columns=data_rows[0])
                else:
                    return None
            except Exception:
                return None
    if df is not None and not df.empty:
        return df
    return None

def load_spreadsheets():
    planilhas_dir = os.path.join(os.path.dirname(__file__), 'planilhas')
    excel_files = glob.glob(os.path.join(planilhas_dir, '*.xlsx'))
    all_data = []
    for file in excel_files:
        file_name = os.path.basename(file)
        temp_df = load_data_from_file(file, file_name)
        if temp_df is not None:
            col_map = {}
            for col in temp_df.columns:
                norm = normalize_column(col)
                if 'num' in norm and 'peca' in norm:
                    col_map[col] = 'Numero da Peca'
                elif 'descricao' in norm:
                    col_map[col] = 'Descricao'
                elif 'quantidade' in norm:
                    col_map[col] = 'Quantidade'
                elif 'localizacao' in norm:
                    col_map[col] = 'Localizacao'
            temp_df = temp_df.rename(columns=col_map)
            required_columns = ['Numero da Peca', 'Descricao', 'Quantidade', 'Localizacao']
            if not all(col in temp_df.columns for col in required_columns):
                continue
            temp_df['Fonte'] = os.path.basename(file)
            temp_df['FonteData'] = extrair_data_planilha(os.path.basename(file))
            temp_df['Numero da Peca Normalizado'] = temp_df['Numero da Peca'].astype(str).str.lower().str.replace(r'\s+', '', regex=True)
            for col in temp_df.columns:
                if temp_df[col].dtype == 'object':
                    temp_df[col] = temp_df[col].astype(str).str.strip()
            all_data.append(temp_df)
    if not all_data:
        return pd.DataFrame()
    df = pd.concat(all_data, ignore_index=True)
    return df

def search_stock(query):
    df = load_spreadsheets()
    if df.empty:
        return []
    query_term = str(query).strip().lower().replace(' ', '')
    if not query_term:
        return []
    results_df = df.copy()
    if 'Numero da Peca Normalizado' in results_df.columns:
        condition_numero_peca = results_df['Numero da Peca Normalizado'].str.contains(query_term, na=False)
    else:
        condition_numero_peca = pd.Series([False] * len(results_df))
    if 'Descricao' in results_df.columns:
        results_df['search_descricao'] = results_df['Descricao'].astype(str).str.lower()
        condition_descricao = results_df['search_descricao'].str.contains(query_term, na=False)
    else:
        condition_descricao = pd.Series([False] * len(results_df))
    if 'Localizacao' in results_df.columns:
        results_df['search_localizacao'] = results_df['Localizacao'].astype(str).str.lower().str.replace(r'\s+', '', regex=True)
        condition_localizacao = results_df['search_localizacao'].str.contains(query_term, na=False)
    else:
        condition_localizacao = pd.Series([False] * len(results_df))
    combined_condition = condition_numero_peca | condition_descricao | condition_localizacao
    results_df = results_df[combined_condition]
    seen = {}
    for _, result in results_df.iterrows():
        key = (result['Numero da Peca'], result['Descricao'], result['Localizacao'])
        fonte_data = result.get('FonteData', '')
        try:
            data_tuple = tuple(map(int, fonte_data.split('/')[::-1]))
        except:
            data_tuple = (0, 0, 0)
        if key not in seen or data_tuple > seen[key][0]:
            seen[key] = (data_tuple, result)
    unique_results = [v[1].to_dict() for v in seen.values()]
    for r in unique_results:
        r['Fonte'] = r.get('FonteData', r.get('Fonte', ''))
    return unique_results[:20]

@app.route('/search', methods=['POST'])
def search():
    data = request.get_json()
    query = data.get('query', '').strip()
    if not query:
        return jsonify({'query': '', 'results': [], 'total': 0, 'message': 'Por favor, digite um termo para buscar'})
    results = search_stock(query)
    response = {
        'query': query,
        'results': results,
        'total': len(results),
        'message': 'Não foi encontrado essa peça nas planilhas.' if len(results) == 0 else f'Encontrados {len(results)} resultados'
    }
    return jsonify(response)

@app.route('/health', methods=['GET'])
def health():
    try:
        df = load_spreadsheets()
        status = {
            'status': 'ok',
            'data_loaded': not df.empty,
            'total_files': len(df["Fonte"].unique()) if not df.empty else 0,
            'total_items': len(df) if not df.empty else 0,
            'columns': df.columns.tolist() if not df.empty else []
        }
        return jsonify(status)
    except Exception as e:
        return jsonify({'status': 'error', 'error': str(e)}), 500

@app.route('/', methods=['GET'])
def index():
    return jsonify({'message': 'API do Sistema de Estoque está ativa.'})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)